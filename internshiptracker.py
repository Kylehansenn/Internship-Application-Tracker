import requests
from bs4 import BeautifulSoup
import datetime
import openpyxl
from openpyxl.styles import Alignment, Side
from openpyxl.styles import PatternFill, Font, Border
import tkinter as tk
from tkinter import *


# https://jobs.thomsonreuters.com/ShowJob/Id/353964/Software-Engineer-Internship-Summer-2020/
# https://jobs.jobvite.com/code42/job/oYc1afw5?__jvst=Career+Site
# https://jobs.lever.co/smartthings/88cf7759-c792-4d20-8cc9-b752f883bd03

# TODO: Divide pullInfo into own function for each item
# TODO: Create location function
# TODO: Exception handling for closing out of excel when open while running
# TODO: Refactor code (Won't do until basically complete with functionality)


def pullCompanyName(info):
    return "could not find"


def pullPositionTitle(info):
    title = ''
    for i in range(len(info)):
        if "Internship" in info[i] or "Intern" in info[i]:
            title = str(info[i]).strip(' ')
            break

    return title


def pullPay(info):
    pay = ''
    for i in range(len(info)):
        if '$' in info[i]:
            pay = str(info[i])
            break

    return pay


def pullDateApplied():
    date = datetime.datetime.now().strftime('%m' + '/' + '%d' + '/' + '%Y')

    return date


def pullCompanyLocation(info):
    location = ''
    cities = open('cities.txt', 'r')

    for i in cities:
        if i in info:
            location = i
            break

    return location



# Pulls HTML
def pullInfo(url):
    soup = BeautifulSoup(requests.get(url).text, features="html.parser")

    # Grabs all information and stores into master list
    info = []
    for item in soup.find_all('title'):
        info.append(item.string)

    for item in soup.find_all('h1'):
        info.append(item.string)

    for item in soup.find_all('h2'):
        info.append(item.string)

    # Images have a separate list for company name
    imgs = soup.find_all('img')
    composs = []
    if len(composs) > 0:
        for image in imgs:
            composs.append(image['alt'])
    else:
        composs.append("Could Not Find")

    # Removes any potential None values
    clean = []
    print(info)
    for val in info:
        if val is not None:
            clean.append(val)
    print(clean)

    # Initialize variables
    company = str(composs[0]).strip('logo')
    title = ""
    date = datetime.datetime.now().strftime('%m' + '/' + '%d' + '/' + '%Y')
    location = "".strip('')

    # Sorts through info and stores information into respected variables
    for i in range(len(clean)):
        if "Internship" in clean[i] or "Intern" in clean[i]:
            title = str(clean[i]).strip(' ')
            break

    tkCompany.insert(0, company)
    tkTitle.insert(0, title)
    tkPay.insert(0, 'Could not find')
    tkLink.insert(0, url)
    tkDateApplied.insert(0, date)
    tkLocation.insert(0, location)


# Function to add internship info to spreadsheet
def updateXl():
    # Opens spreadsheet
    file = 'Internship Tracker.xlsx'
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    # Style formatting
    bg = PatternFill(start_color='595959',
                          end_color='595959',
                          fill_type='solid')

    border = Border(
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin'))

    font = Font(color='ffffff')

    # Imports data into cells
    ws.insert_rows(3)

    # Insert company
    ws['A3'] = tkCompany.get()
    ws['A3'].fill = bg

    # Insert role title
    ws['B3'] = tkTitle.get()

    # Insert pay
    ws['C3'] = 'N/A'

    # Insert URL
    ws['D3'] = tkLink.get()

    # Insert date applied
    ws['E3'] = tkDateApplied.get()

    # Insert location
    ws['F3'] = tkLocation.get()

    # Blank for drop down menus
    # TODO: Change from blank to default values
    ws['G3'] = ''
    ws['H3'] = ''
    ws['I3'] = ''
    ws['J3'] = ''

    # Applying style formats to cells
    ws['C3'].alignment = Alignment(horizontal='center')
    ws['E3'].alignment = Alignment(horizontal='right')
    ws['E3'].font = Font(bold=True)
    ws['A3'].fill = ws['B3'].fill = ws['C3'].fill = ws['D3'].fill = ws['E3'].fill = ws['F3'].fill = ws['G3'].fill\
        = ws['H3'].fill = ws['I3'].fill = ws['J3'].fill = bg
    ws['A3'].border = ws['B3'].border = ws['C3'].border = ws['D3'].border = ws['E3'].border = ws['F3'].border\
        = ws['G3'].border = ws['H3'].border = ws['I3'].border = ws['J3'].border = border
    ws['A3'].font = ws['B3'].font = ws['C3'].font = ws['D3'].font = ws['E3'].font = ws['F3'].font \
        = ws['G3'].font = ws['H3'].font = ws['I3'].font = ws['J3'].font = font

    wb.save(file)


def start():
    # Webscraper converting HTMl into readable text
    soup = BeautifulSoup(requests.get(urlBox.get()).text, features="html.parser")

    # Looks through page source for potential information
    info = []
    for item in soup.find_all('title'):
        info.append(item.string)

    for item in soup.find_all('h1'):
        info.append(item.string)

    for item in soup.find_all('h2'):
        info.append(item.string)

    # Call all functions for finding respected item
    #pullInfo(urlBox.get())

    tkCompany.insert(0, pullCompanyName(info))
    tkTitle.insert(0, pullPositionTitle(info))
    tkPay.insert(0, pullPay(info))
    tkLink.insert(0, urlBox.get())
    tkDateApplied.insert(0, pullDateApplied())
    tkLocation.insert(0, pullCompanyLocation(info))


root = tk.Tk()
root.configure(bg='#31393C')

# URL Input Entry
Label(root, text='Enter job description URL:', font=("Helvetica", 16, 'bold'),
      fg='#94DDBC', bg='#31393C').grid(row=0, column=0)
urlBox = tk.Entry(root)
urlBox.grid(row=0, column=1)
submit = tk.Button(root, text='Submit', command=start).grid(row=0, column=2)

# Company Name
Label(root, text='Company:', font=("Helvetica", 10, 'bold'),  bg='#31393C', fg='#94DDBC').grid(row=1, column=0)
tkCompany = tk.Entry(root)
tkCompany.grid(row=1, column=1)

# Role Title
Label(root, text='Position Title:', font=("Helvetica", 10, 'bold'),  bg='#31393C', fg='#94DDBC').grid(row=2, column=0)
tkTitle = tk.Entry(root)
tkTitle.grid(row=2, column=1)

# Pay
Label(root, text='Pay:', font=("Helvetica", 10, 'bold'),  bg='#31393C', fg='#94DDBC').grid(row=3, column=0)
tkPay = tk.Entry(root)
tkPay.grid(row=3, column=1)

# Link to advert
Label(root, text='Advert Link:', font=("Helvetica", 10, 'bold'),  bg='#31393C' , fg='#94DDBC').grid(row=4, column=0)
tkLink = Entry(root)
tkLink.grid(row=4, column=1)

# Date Applied
Label(root, text='Date Applied:', font=("Helvetica", 10, 'bold'), fg='#94DDBC', bg='#31393C').grid(row=5, column=0)
tkDateApplied = Entry(root)
tkDateApplied.grid(row=5, column=1)

# Location
Label(root, text='Location:', font=("Helvetica", 10, 'bold'),  bg='#31393C' , fg='#94DDBC').grid(row=6, column=0)
tkLocation = tk.Entry(root)
tkLocation.grid(row=6, column=1)

# Update Excel
tkUpdate = Button(root, text='Update Excel Spreadsheet', command=updateXl)
tkUpdate.grid(row=7, column=1)

root.mainloop()




